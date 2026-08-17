from __future__ import annotations

import io
import math
import re
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment


APP_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = APP_DIR / "template.xlsx"
TARGET_FACULTY = "数学与统计学院"
EMPTY_MARKERS = {"", "nan", "none", "无", "暂无", "不清楚", "未填写"}


@dataclass(frozen=True)
class Columns:
    week: str
    student_name: str
    major: str
    class_name: str
    teacher_faculty: str
    teacher: str
    course: str
    feedback: str


def text(value: object) -> str:
    """Convert a worksheet value to clean text without leaking pandas NaN."""
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def compact_header(value: object) -> str:
    return re.sub(r"[\s\n\r]+", "", text(value))


def find_column(
    columns: Iterable[object],
    required: tuple[str, ...],
    excluded: tuple[str, ...] = (),
    preferred: tuple[str, ...] = (),
) -> str:
    candidates: list[tuple[int, str]] = []
    for original in columns:
        normalized = compact_header(original)
        if all(token in normalized for token in required) and not any(token in normalized for token in excluded):
            score = sum(token in normalized for token in preferred)
            candidates.append((score, str(original)))
    if not candidates:
        raise ValueError(f"未找到列：{' + '.join(required)}。请确认上传的是原始教学反馈表。")
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def detect_columns(df: pd.DataFrame) -> Columns:
    columns = list(df.columns)
    return Columns(
        week=find_column(columns, ("周数",)),
        student_name=find_column(columns, ("姓名",), preferred=("您的信息",), excluded=("教师信息",)),
        major=find_column(columns, ("专业",), preferred=("3、",), excluded=("教师",)),
        class_name=find_column(columns, ("班级",)),
        teacher_faculty=find_column(columns, ("教师所在院",)),
        teacher=find_column(columns, ("教师信息", "姓名")),
        course=find_column(columns, ("教授课程",)),
        feedback=find_column(columns, ("反馈内容",)),
    )


def detect_week(df: pd.DataFrame, week_column: str, filename: str) -> int:
    filename_match = re.search(r"第?\s*(\d{1,2})\s*周", filename)
    values = pd.to_numeric(
        df[week_column].astype(str).str.extract(r"(\d{1,2})", expand=False),
        errors="coerce",
    ).dropna().astype(int)
    if values.empty:
        raise ValueError("周数列中没有可识别的数字。")
    if filename_match:
        named_week = int(filename_match.group(1))
        if named_week in set(values):
            return named_week
    counts = Counter(values.tolist())
    max_count = max(counts.values())
    return max(week for week, count in counts.items() if count == max_count)


def normalize_grade(feedback: str, class_name: str) -> str:
    match = re.search(r"(?:20)?(\d{2})\s*级", feedback)
    if match:
        return match.group(1)
    match = re.search(r"(?<!\d)(\d{2})(?:\d{2}|级)", class_name)
    return match.group(1) if match else "未知"


def normalize_major(value: str) -> str:
    """Normalize only harmless suffix variants; never collapse different majors."""
    value = re.sub(r"\s+", "", value)
    value = re.sub(r"(?:专业)?类专业$", "类", value)
    value = re.sub(r"专业$", "", value)
    value = {
        "材料": "材料类",
        "材料类": "材料类",
        "数学": "数学类",
        "数学类": "数学类",
    }.get(value, value)
    return value or "未注明专业"


def major_from_feedback(feedback: str, fallback: str) -> str:
    """Prefer the class/major named in the feedback over the reporter's own major."""
    match = re.search(
        r"(?:20)?\d{2}(?:级)?\s*(.+?)(?=(?:的)?(?:同学|学生)|有同学)",
        feedback,
    )
    if match:
        candidate = re.sub(r"[／/](?:类|专业)$", "", match.group(1).strip())
        if candidate:
            return normalize_major(candidate)
    return normalize_major(fallback)


def major_in_sentence(major: str) -> str:
    if major.endswith(("专业", "类")):
        return major
    return f"{major}专业"


def normalize_course(value: str) -> str:
    value = text(value)
    value = re.sub(r"^[《〈\"“']+|[》〉\"”']+$", "", value).strip()
    value = value.replace("《", "").replace("》", "").replace("〈", "").replace("〉", "")
    value = {
        "高等数学（一）B": "高等数学B（一）",
        "高等数学（二）B": "高等数学B（二）",
    }.get(value, value)
    return value or "未注明课程"


def remove_personal_information(value: str, student_name: str) -> str:
    value = re.sub(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", "", value)
    value = re.sub(r"(?<!\d)1[3-9]\d{9}(?!\d)", "", value)
    value = re.sub(r"(?<!\d)\d{8,12}(?!\d)", "", value)
    value = re.sub(r"(?:学号|姓名|联系方式|手机号)\s*[:：]?\s*[\w\u4e00-\u9fff-]+", "", value)
    if student_name:
        escaped = re.escape(student_name)
        value = re.sub(rf"{escaped}\s*(?:同学|信息员)?", "", value)
    return value


def strip_existing_lead(value: str, course: str) -> str:
    marker = re.search(r"(?:同学|学生)?(?:反映|反馈)[，,：:\s]*", value)
    if marker:
        value = value[marker.end():]

    course_pattern = re.escape(course)
    patterns = [
        rf"^(?:老师)?\s*(?:在)?\s*(?:所授|所教授|教授|讲授)的?\s*[《〈]?{course_pattern}[》〉]?\s*(?:课程)?\s*(?:中)?\s*[，,：:\s]*",
        r"^(?:老师)?\s*(?:在)?\s*(?:所授|所教授|教授|讲授)的?\s*[《〈]?[^》〉，,。]{1,40}?[》〉]?\s*(?:课程中|课程)\s*[，,：:\s]*",
        rf"^(?:老师)?\s*(?:在)?\s*[《〈]?{course_pattern}[》〉]?\s*(?:课程)?\s*(?:中)\s*[，,：:\s]*",
        rf"^[《〈]?{course_pattern}[》〉]?\s*(?:课程)?\s*(?:中)?\s*[，,：:\s]*",
    ]
    for pattern in patterns:
        updated = re.sub(pattern, "", value, count=1)
        if updated != value:
            value = updated
            break
    return value


def clean_feedback(raw: str, student_name: str, teacher: str, course: str) -> str:
    value = text(raw).replace("反应", "反映")
    value = remove_personal_information(value, student_name)
    if teacher:
        value = re.sub(rf"{re.escape(teacher)}\s*老师", "老师", value)
        if len(teacher) >= 2:
            value = re.sub(rf"{re.escape(teacher[0])}\s*老师", "老师", value)
    value = strip_existing_lead(value, course)
    value = re.sub(r"(?:有)?同学(?:们)?(?:反映|反馈)[，,：:\s]*", "", value)
    value = value.replace("反应", "反映")
    value = re.sub(r"\s*\n\s*", "", value)
    value = re.sub(r"\s+", "", value)
    value = re.sub(r"^[，,。；;：:\s]+", "", value)
    value = re.sub(r"[。！？!?；;，,\s]+$", "", value)
    return value


def is_empty_feedback(value: str) -> bool:
    normalized = re.sub(r"[\s，,。.!！?？；;：:]+", "", text(value)).lower()
    return normalized in {
        "", "无", "暂无", "没有", "无反馈", "暂无反馈", "无意见", "无问题",
        "本周无反馈", "本周暂无反馈",
    }


def sentence_for_group(grade: str, major: str, course_contents: dict[str, list[str]]) -> str:
    course_clauses: list[str] = []
    for course, contents in course_contents.items():
        unique_contents = list(dict.fromkeys(item for item in contents if item))
        joined = "；".join(unique_contents)
        course_clauses.append(f"老师在所授的《{course}》课程中，{joined}")
    lead = f"{grade}级{major_in_sentence(major)}的同学反映，"
    return lead + "；".join(course_clauses) + "。"


def consolidate_course_contents(course_contents: dict[str, list[str]]) -> dict[str, list[str]]:
    """Merge numbered spelling variants when the unnumbered course also exists."""
    consolidated: dict[str, list[str]] = {}
    names = list(course_contents)
    for course, contents in course_contents.items():
        base = re.sub(r"(?:[（(]?[一二三四五六七八九十\d]+[）)]?)$", "", course).strip()
        target = base if base and base in names else course
        consolidated.setdefault(target, []).extend(contents)
    return consolidated


def transform_feedback(data: bytes, filename: str) -> tuple[pd.DataFrame, dict[str, object]]:
    df = pd.read_excel(io.BytesIO(data), dtype=str, engine="openpyxl")
    if df.empty:
        raise ValueError("上传的工作簿没有数据。")
    columns = detect_columns(df)
    week = detect_week(df, columns.week, filename)

    week_numbers = pd.to_numeric(
        df[columns.week].astype(str).str.extract(r"(\d{1,2})", expand=False),
        errors="coerce",
    )
    # The uploaded workbook is the processing boundary. Some weekly exports contain
    # late/early submissions whose recorded week differs from the filename; those
    # rows still belong in the requested output and must not be silently discarded.
    current = df.copy()
    faculty_mask = current[columns.teacher_faculty].fillna("").astype(str).str.contains(TARGET_FACULTY, regex=False)
    faculty_rows = current.loc[faculty_mask].copy()

    records: list[dict[str, str]] = []
    skipped_empty = 0
    for _, row in faculty_rows.iterrows():
        teacher = text(row[columns.teacher])
        course = normalize_course(row[columns.course])
        raw_feedback = text(row[columns.feedback])
        if teacher.lower() in EMPTY_MARKERS or course.lower() in EMPTY_MARKERS or is_empty_feedback(raw_feedback):
            skipped_empty += 1
            continue
        student_name = text(row[columns.student_name])
        major = major_from_feedback(raw_feedback, text(row[columns.major]))
        grade = normalize_grade(raw_feedback, text(row[columns.class_name]))
        content = clean_feedback(raw_feedback, student_name, teacher, course)
        if not content:
            skipped_empty += 1
            continue
        records.append({
            "年级": grade,
            "专业": major,
            "教师": teacher,
            "课程": course,
            "内容": content,
        })

    if not records:
        raise ValueError("上传文件中没有找到数学与统计学院教师的有效反馈。")

    grouped: dict[tuple[str, str, str], dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for record in records:
        # 专业是不可省略的分组维度。同一教师面向“数学类”和
        # “数学与应用数学”授课时，必须生成两条独立记录。
        key = (record["年级"], record["专业"], record["教师"])
        grouped[key][record["课程"]].append(record["内容"])

    def grade_key(value: str) -> int:
        return int(value) if value.isdigit() else -1

    output_rows = []
    for (grade, major, teacher), course_contents in sorted(
        grouped.items(), key=lambda item: (-grade_key(item[0][0]), item[0][1], item[0][2])
    ):
        course_contents = consolidate_course_contents(course_contents)
        courses = "、".join(course_contents.keys())
        output_rows.append({
            "年级": f"{grade}级" if grade.isdigit() else grade,
            "专业": major,
            "教师": teacher,
            "课程": courses,
            "反馈信息": sentence_for_group(grade, major, course_contents),
        })

    result = pd.DataFrame(output_rows)
    source_group_counts = Counter(
        (record["年级"], record["专业"], record["教师"])
        for record in records
    )
    expected_groups = set(source_group_counts)
    actual_groups = {
        (
            re.sub(r"级$", "", text(row["年级"])),
            normalize_major(text(row["专业"])),
            text(row["教师"]),
        )
        for row in output_rows
    }
    if actual_groups != expected_groups:
        missing = sorted(expected_groups - actual_groups)
        raise RuntimeError(f"专业分组校验失败，以下教师记录未被保留：{missing}")

    teacher_majors: dict[str, set[str]] = defaultdict(set)
    for _, major, teacher in expected_groups:
        teacher_majors[teacher].add(major)
    multi_major_teachers = {
        teacher: sorted(majors)
        for teacher, majors in teacher_majors.items()
        if len(majors) > 1
    }
    summary = {
        "week": week,
        "source_rows": len(df),
        "week_rows": len(current),
        "faculty_rows": len(faculty_rows),
        "valid_rows": len(records),
        "skipped_empty": skipped_empty,
        "merged_rows": len(result),
        "grades": result["年级"].drop_duplicates().tolist(),
        "detected_weeks": sorted({int(value) for value in week_numbers.dropna().tolist()}),
        "multi_major_teachers": multi_major_teachers,
        "source_groups": [
            {
                "年级": f"{grade}级" if grade.isdigit() else grade,
                "专业": major,
                "教师": teacher,
                "有效原始记录数": count,
            }
            for (grade, major, teacher), count in sorted(
                source_group_counts.items(),
                key=lambda item: (-grade_key(item[0][0]), item[0][1], item[0][2]),
            )
        ],
    }
    return result, summary


def chinese_number(number: int) -> str:
    digits = "零一二三四五六七八九"
    if number < 10:
        return digits[number]
    if number < 20:
        return "十" + (digits[number % 10] if number % 10 else "")
    if number < 100:
        return digits[number // 10] + "十" + (digits[number % 10] if number % 10 else "")
    return str(number)


def estimate_row_height(feedback: str) -> float:
    visual_length = sum(2 if ord(char) > 127 else 1 for char in feedback)
    lines = max(2, math.ceil(visual_length / 105))
    return min(165.0, max(36.0, 15.0 * (lines + 1)))


def write_template(result: pd.DataFrame, week: int) -> bytes:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError("项目缺少 template.xlsx，请将模板文件放在 app.py 同目录。")
    workbook = load_workbook(TEMPLATE_PATH)
    worksheet = workbook[workbook.sheetnames[0]]

    style_sources = {column: copy(worksheet.cell(3, column)._style) for column in range(1, 6)}
    # A3 belongs to a vertically merged range and therefore has no bottom border.
    # Use the adjacent body-cell style so newly sized grade groups keep a complete outline.
    style_sources[1] = copy(worksheet.cell(3, 2)._style)
    number_formats = {column: worksheet.cell(3, column).number_format for column in range(1, 6)}
    protections = {column: copy(worksheet.cell(3, column).protection) for column in range(1, 6)}

    for merged in list(worksheet.merged_cells.ranges):
        if merged.min_row >= 3:
            worksheet.unmerge_cells(str(merged))

    if worksheet.max_row >= 3:
        worksheet.delete_rows(3, worksheet.max_row - 2)
    worksheet["A1"] = f"关于数学与统计学院教师第{chinese_number(week)}周教学信息反馈"

    for row_index, record in enumerate(result.to_dict("records"), start=3):
        values = [record["年级"], record["专业"], record["教师"], record["课程"], record["反馈信息"]]
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column_index, value)
            cell._style = copy(style_sources[column_index])
            cell.number_format = number_formats[column_index]
            cell.protection = copy(protections[column_index])
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=column_index == 5,
                shrink_to_fit=False,
            )
        worksheet.row_dimensions[row_index].height = estimate_row_height(record["反馈信息"])

    start = 3
    while start <= len(result) + 2:
        grade = worksheet.cell(start, 1).value
        end = start
        while end + 1 <= len(result) + 2 and worksheet.cell(end + 1, 1).value == grade:
            end += 1
        if end > start:
            worksheet.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        worksheet.cell(start, 1).alignment = Alignment(horizontal="center", vertical="center")
        start = end + 1

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def output_filename(source_filename: str, fallback_week: int) -> str:
    """Build a canonical output name, taking the week from the uploaded filename."""
    stem = Path(source_filename).stem
    stem = re.sub(r"（\d+）|\(\d+\)|优化版|修正版|副本", "", stem)

    numeric_match = re.search(r"第\s*(\d{1,2})\s*周", stem)
    if numeric_match:
        week = int(numeric_match.group(1))
    else:
        chinese_match = re.search(r"第\s*([一二三四五六七八九十]{1,3})\s*周", stem)
        if chinese_match:
            value = chinese_match.group(1)
            digits = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
                      "六": 6, "七": 7, "八": 8, "九": 9}
            if "十" in value:
                left, right = value.split("十", 1)
                week = (digits.get(left, 1) if left else 1) * 10 + digits.get(right, 0)
            else:
                week = digits[value]
        else:
            week = fallback_week

    return f"关于数学与统计学院第{chinese_number(week)}周教学信息反馈.xlsx"


def main() -> None:
    st.set_page_config(page_title="教学信息反馈生成器", page_icon="📘", layout="wide")
    st.markdown(
        """
        <style>
        .block-container {max-width: 1120px; padding-top: 2.4rem;}
        [data-testid="stFileUploaderDropzone"] {border: 1px dashed #315f56; background: #f6f8f5;}
        .tool-note {color:#60706b; line-height:1.8; margin-bottom:1.2rem;}
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.title("教学信息反馈生成器")
    st.markdown(
        '<div class="tool-note">上传每周原始反馈表，系统将自动识别最新周次、筛选数学与统计学院教师、合并同类反馈，并套用学院标准模板。</div>',
        unsafe_allow_html=True,
    )

    uploaded = st.file_uploader("上传每周教学反馈表", type=["xlsx"], help="上传问卷平台导出的原始 .xlsx 文件")
    if uploaded is None:
        st.info("请选择一个 Excel 文件。处理仅在当前运行环境中进行。")
        with st.expander("处理规则"):
            st.write("自动筛选教师所在院为“数学与统计学院”的记录，并按年级降序排列。")
            st.write("同年级、同专业、同教师的反馈合并；不同年级或专业保持分开。")
            st.write("自动删除信息员姓名与常见个人信息，统一使用“反映”和中文句号。")
        return

    try:
        source_bytes = uploaded.getvalue()
        result, summary = transform_feedback(source_bytes, uploaded.name)
        output_bytes = write_template(result, int(summary["week"]))
    except Exception as exc:
        st.error(f"处理失败：{exc}")
        st.stop()

    st.success(f"已完成第 {summary['week']} 周反馈整理")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("上传记录", summary["week_rows"])
    col2.metric("数统学院记录", summary["faculty_rows"])
    col3.metric("有效反馈", summary["valid_rows"])
    col4.metric("合并后条目", summary["merged_rows"])
    if summary["skipped_empty"]:
        st.caption(f"已跳过 {summary['skipped_empty']} 条“无”、空内容或只有标准开头的反馈。")
    if len(summary["detected_weeks"]) > 1:
        weeks = "、".join(str(item) for item in summary["detected_weeks"])
        st.caption(f"文件中检测到周次 {weeks}；按整份上传文件处理，标题周次取文件名中的第 {summary['week']} 周。")
    if summary["multi_major_teachers"]:
        details = "；".join(
            f"{teacher}（{'、'.join(majors)}）"
            for teacher, majors in summary["multi_major_teachers"].items()
        )
        st.info(f"已按专业分别保留同一教师的反馈：{details}")

    with st.expander("源记录分组核对"):
        st.caption("这里显示生成前的有效记录数；源表没有的教师—专业组合不会由程序推测或补写。")
        st.dataframe(
            pd.DataFrame(summary["source_groups"]),
            use_container_width=True,
            hide_index=True,
        )

    st.subheader("结果预览")
    st.dataframe(result, use_container_width=True, hide_index=True, height=min(620, 72 + len(result) * 36))
    st.download_button(
        "下载教学信息反馈表",
        data=output_bytes,
        file_name=output_filename(uploaded.name, int(summary["week"])),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
    st.caption("下载文件已完整保留模板的标题样式、边框、列宽、合并单元格、页面设置和打印布局。")


if __name__ == "__main__":
    main()
